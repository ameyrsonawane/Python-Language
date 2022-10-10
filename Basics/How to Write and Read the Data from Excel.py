# How to write the data in Excel -->

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

#ws['A1'] = "How to read the data from excel using openpyxl module"

testdata = [["Amey","Pune"],["Shraddha","Jalgaon"],["Nishtha","Pune"]]
for data in testdata:
    ws.append(data)
    wb.save("demo_excel.xlsx")

# How to read the data from Excel -->

from openpyxl import workbook, load_workbook
wb = load_workbook(filename = "demo_excel.xlsx")
sh = wb.active
print(sh['A3'].value)

row = sh.max_row
column = sh.max_column

# How to read complete data from excel -->

print("Row_count-",row)
print("Column_count-",column)

for i in range(1, row+1):
    for j in range(1, column+1):
        print(sh.cell(row = i, column = j).value, end=' ')
    print("\n")
